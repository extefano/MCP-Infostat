from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from mcp_infostat.config import AppConfig
from mcp_infostat.errors import InfoStatError
from mcp_infostat.ui.launcher import InfoStatLauncher


@dataclass
class DatasetInfo:
    path: Path
    rows: int
    columns: list[str]
    has_header: bool
    delimiter: str | None
    file_format: str
    sheet_name: str | None = None


@dataclass
class InfoStatSession:
    pid: int | None = None
    ui_backend: str | None = None
    is_ready: bool = False
    active_dataset: Path | None = None
    dataset_info: DatasetInfo | None = None
    results_buffer: list[str] = field(default_factory=list)


class InfoStatSessionManager:
    SUPPORTED_DATA_EXTENSIONS = {".csv", ".txt", ".xls", ".xlsx", ".dbf"}

    def __init__(self, config: AppConfig, launcher: InfoStatLauncher | None = None):
        self.config = config
        self.launcher = launcher or InfoStatLauncher(config.infostat.exe_path)
        self.session = InfoStatSession()

    def launch(self, exe_path: str | None = None, timeout: float | None = None) -> dict[str, Any]:
        if self.session.is_ready:
            raise InfoStatError(
                code="SESSION_ALREADY_ACTIVE",
                message="InfoStat ya se encuentra en ejecucion.",
                details={"pid": self.session.pid},
            )

        app_info = self.launcher.launch(
            exe_path=exe_path,
            timeout=timeout or self.config.timeouts.launch_seconds,
        )
        self.session.pid = app_info["pid"]
        self.session.ui_backend = app_info.get("backend")
        self.session.is_ready = True
        return {
            "running": True,
            "pid": self.session.pid,
            "ui_backend": self.session.ui_backend,
            "version": self.config.infostat.version,
            "dataset_loaded": False,
        }

    def status(self) -> dict[str, Any]:
        running = self.session.is_ready and self.launcher.is_ready()
        self.session.is_ready = running
        return {
            "running": running,
            "dataset_loaded": self.session.active_dataset is not None,
            "dataset_name": self.session.active_dataset.name if self.session.active_dataset else None,
            "ready": running,
            "ui_backend": self.session.ui_backend,
            "version": self.config.infostat.version,
        }

    def close(self, save_before_close: bool = False) -> dict[str, Any]:
        if not self.session.is_ready:
            raise InfoStatError(
                code="SESSION_NOT_ACTIVE",
                message="No hay una sesion activa para cerrar.",
            )

        self.launcher.close(save_before_close=save_before_close)
        had_unsaved_changes = False

        self.session = InfoStatSession()
        return {
            "running": False,
            "unsaved_changes": had_unsaved_changes,
        }

    def data_load(
        self,
        file_path: Path,
        sheet_name: str | None,
        delimiter: str | None,
        has_header: bool,
    ) -> dict[str, Any]:
        self._ensure_ready()

        ext = file_path.suffix.lower()
        if ext not in self.SUPPORTED_DATA_EXTENSIONS:
            raise InfoStatError(
                code="DATA_FORMAT_NOT_IMPLEMENTED",
                message="Formato de datos no soportado para carga en InfoStat.",
                details={
                    "requested_extension": ext,
                    "supported_extensions": sorted(self.SUPPORTED_DATA_EXTENSIONS),
                },
            )

        rows: int
        columns: list[str]
        separator: str | None = None
        effective_sheet: str | None = None

        if ext in {".csv", ".txt"}:
            separator = delimiter or self._detect_delimiter(file_path)
            rows, columns = self._scan_csv(file_path, delimiter=separator, has_header=has_header)
        elif ext in {".xls", ".xlsx"}:
            rows, columns, effective_sheet = self._scan_excel(
                file_path=file_path,
                extension=ext,
                sheet_name=sheet_name,
                has_header=has_header,
            )
        else:
            rows, columns = self._scan_dbf(file_path=file_path)

        self.session.active_dataset = file_path
        self.session.dataset_info = DatasetInfo(
            path=file_path,
            rows=rows,
            columns=columns,
            has_header=has_header,
            delimiter=separator,
            file_format=ext,
            sheet_name=effective_sheet,
        )

        try:
            ui_loaded = bool(self.launcher.load_file_via_keyboard(file_path=file_path))
        except InfoStatError:
            raise
        except Exception as exc:
            raise InfoStatError(
                code="DATA_LOAD_UI_FAILED",
                message="Error no esperado al cargar el archivo en la UI de InfoStat.",
                details={
                    "file_path": str(file_path),
                    "exception_type": type(exc).__name__,
                    "raw": str(exc),
                },
            ) from exc

        if not ui_loaded:
            raise InfoStatError(
                code="DATA_LOAD_UI_FAILED",
                message="No se pudo cargar el archivo en InfoStat mediante estrategia de teclado.",
                details={"file_path": str(file_path)},
            )

        return {
            "file_path": str(file_path),
            "rows": rows,
            "cols": len(columns),
            "columns": columns,
            "sheet_name": effective_sheet,
            "delimiter": separator,
            "has_header": has_header,
            "format": ext,
            "mode": "ui_keyboard",
            "ui_loaded": True,
            "ui_backend": self.session.ui_backend,
        }

    def data_get_info(self) -> dict[str, Any]:
        self._ensure_ready()
        info = self.session.dataset_info
        if info is None:
            raise InfoStatError(
                code="DATASET_NOT_LOADED",
                message="No hay un dataset cargado en la sesion.",
            )

        return {
            "file_path": str(info.path),
            "n_rows": info.rows,
            "n_cols": len(info.columns),
            "columns": info.columns,
            "delimiter": info.delimiter,
            "has_header": info.has_header,
            "format": info.file_format,
            "sheet_name": info.sheet_name,
        }

    def append_result(self, text: str) -> None:
        self.session.results_buffer.append(text)

    def get_last_result(self) -> str:
        if not self.session.results_buffer:
            return ""
        return self.session.results_buffer[-1]

    def _ensure_ready(self) -> None:
        if not self.session.is_ready or not self.launcher.is_ready():
            self.session.is_ready = False
            raise InfoStatError(
                code="SESSION_NOT_ACTIVE",
                message="InfoStat no esta activo. Llama a infostat_launch primero.",
            )

    @staticmethod
    def _detect_delimiter(file_path: Path) -> str:
        sample = file_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        first_line = sample[0] if sample else ""
        if ";" in first_line and "," not in first_line:
            return ";"
        return ","

    @staticmethod
    def _scan_csv(file_path: Path, delimiter: str, has_header: bool) -> tuple[int, list[str]]:
        with file_path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = list(reader)

        if not rows:
            return 0, []

        if has_header:
            header = rows[0]
            data_rows = rows[1:]
            return len(data_rows), header

        width = len(rows[0])
        columns = [f"col_{i+1}" for i in range(width)]
        return len(rows), columns

    @classmethod
    def _scan_excel(
        cls,
        file_path: Path,
        extension: str,
        sheet_name: str | None,
        has_header: bool,
    ) -> tuple[int, list[str], str]:
        if extension == ".xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as exc:  # pragma: no cover - depende de entorno
                raise InfoStatError(
                    code="DEPENDENCY_MISSING",
                    message="Falta dependencia para leer archivos XLSX.",
                    details={"dependency": "openpyxl", "extension": extension},
                ) from exc

            workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
            try:
                if sheet_name:
                    if sheet_name not in workbook.sheetnames:
                        raise InfoStatError(
                            code="DATA_SHEET_NOT_FOUND",
                            message="No se encontro la hoja solicitada en el archivo Excel.",
                            details={"sheet_name": sheet_name, "available_sheets": workbook.sheetnames},
                        )
                    selected_sheet_name = sheet_name
                else:
                    if not workbook.sheetnames:
                        return 0, [], ""
                    selected_sheet_name = workbook.sheetnames[0]

                worksheet = workbook[selected_sheet_name]
                rows = []
                for raw_row in worksheet.iter_rows(values_only=True):
                    values = cls._normalize_row_values(raw_row)
                    if values:
                        rows.append(values)
            finally:
                workbook.close()

            row_count, columns = cls._rows_to_metadata(rows=rows, has_header=has_header)
            return row_count, columns, selected_sheet_name

        try:
            import xlrd
        except Exception as exc:  # pragma: no cover - depende de entorno
            raise InfoStatError(
                code="DEPENDENCY_MISSING",
                message="Falta dependencia para leer archivos XLS.",
                details={"dependency": "xlrd", "extension": extension},
            ) from exc

        workbook = xlrd.open_workbook(str(file_path))
        if sheet_name:
            try:
                worksheet = workbook.sheet_by_name(sheet_name)
                selected_sheet_name = sheet_name
            except Exception as exc:
                raise InfoStatError(
                    code="DATA_SHEET_NOT_FOUND",
                    message="No se encontro la hoja solicitada en el archivo Excel.",
                    details={"sheet_name": sheet_name, "available_sheets": workbook.sheet_names()},
                ) from exc
        else:
            if workbook.nsheets == 0:
                return 0, [], ""
            worksheet = workbook.sheet_by_index(0)
            selected_sheet_name = worksheet.name

        rows = []
        for index in range(worksheet.nrows):
            values = cls._normalize_row_values(worksheet.row_values(index))
            if values:
                rows.append(values)

        row_count, columns = cls._rows_to_metadata(rows=rows, has_header=has_header)
        return row_count, columns, selected_sheet_name

    @staticmethod
    def _scan_dbf(file_path: Path) -> tuple[int, list[str]]:
        try:
            from dbfread import DBF
        except Exception as exc:  # pragma: no cover - depende de entorno
            raise InfoStatError(
                code="DEPENDENCY_MISSING",
                message="Falta dependencia para leer archivos DBF.",
                details={"dependency": "dbfread", "extension": ".dbf"},
            ) from exc

        table = DBF(str(file_path), load=True, char_decode_errors="ignore")
        columns = [field.name for field in table.fields]
        rows = len(table.records)
        return rows, columns

    @staticmethod
    def _normalize_row_values(raw_row: Any) -> list[Any]:
        values = list(raw_row)
        while values and (values[-1] is None or str(values[-1]).strip() == ""):
            values.pop()
        return values

    @classmethod
    def _rows_to_metadata(cls, rows: list[list[Any]], has_header: bool) -> tuple[int, list[str]]:
        if not rows:
            return 0, []

        first_row = rows[0]
        if has_header:
            header = cls._normalize_header(first_row)
            return max(len(rows) - 1, 0), header

        width = len(first_row)
        return len(rows), [f"col_{i+1}" for i in range(width)]

    @staticmethod
    def _normalize_header(values: list[Any]) -> list[str]:
        normalized: list[str] = []
        for index, value in enumerate(values):
            name = str(value).strip() if value is not None else ""
            normalized.append(name or f"col_{index+1}")
        return normalized
